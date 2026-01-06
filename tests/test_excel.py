import datetime

import config.jp_data as jp
from openpyxl import load_workbook
import utils.excel_operations as excel

goods_id = "e4j7l18e01"
platform = "淘宝"
tb_dict = {
    "tb_time_index": 23
}
lm_name = "涉及/x/t"


# 数据返填回表格
def write_csv(goods_id, platform, tb_dict, lm_name):
    # 获取表格主路径
    excel_path = "津铺商品库.xlsm"

    # 准备商品数据列表
    index_list = []

    # 读取全部数据, 二维列表
    wb = load_workbook(excel_path)
    ws = wb.active

    # 获取分类和发布时间所在列
    sort_index = 0
    time_index = 0
    if platform == '淘宝':
        sort_index = tb_dict["tb_time_index"]
        time_index = sort_index + 1
    elif platform == '闲鱼':
        ws['J' + str(index_list[0] + 1)] = '已发布'
    elif platform == '微店':
        ws['K' + str(index_list[0] + 1)] = '已发布'

    # 获取当前日期时间
    today_date = str(datetime.date.today())

    print(today_date, type(today_date))

    for i, row in enumerate(ws.values):
        if row[0] == goods_id:
            index_list.append(i)
            print(i+1, sort_index)
            print(i+1, time_index)
            ws.cell(row=i+1, column=sort_index).value = lm_name
            ws.cell(row=i+1, column=time_index).value = today_date

    wb = load_workbook(excel_path, keep_vba=True)
    wb.save("津铺商品库.xlsm")
    # wb.close()

excel_path = r"C:\Users\mac\Desktop\津铺商品库 -  V1.xlsm"
tb_dict_list, xy_dict_list, wd_dict_list = excel.read_csv_to_dict(excel_path)

for i in tb_dict_list:
    print(i)