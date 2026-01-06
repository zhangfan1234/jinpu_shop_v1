import time

from openpyxl import load_workbook
import config.jp_data as jp
import datetime


# 读取表格数据并处理成字典列表
def read_csv_to_dict(excel_path):
    # 准备商品数据字典列表
    goods_dict_list = []

    # 读取全部数据, 二维列表
    wb = load_workbook(excel_path)
    ws = wb.active

    # 获取指定字段的索引
    title_list = []
    for cell in ws[1]:
        title_list.append(cell.value)

    id_index = title_list.index('编号')
    name_index = title_list.index('标题')
    desc_index = title_list.index('详情描述')
    brand_index = title_list.index('品牌')
    sku_price_index = title_list.index('sku_颜色_价格')
    sku_code_index = title_list.index('sku_编码')
    # sku_code_index = title_list.index('sku_库存编码')
    price_index = title_list.index('价格')
    xy_mendian_index = title_list.index("闲鱼门店")
    stock_index = title_list.index('sku_库存')
    # stock_index = title_list.index('sku_库存编码')
    tb_time_index = title_list.index('淘宝发布时间')
    xy_time_index = title_list.index('闲鱼发布时间')
    wd_time_index = title_list.index('微店发布时间')

    # 判断数据写入到字典列表
    tb_dict_list = []
    xy_dict_list = []
    wd_dict_list = []
    for row in ws.iter_rows(min_row=2, values_only=True):

        # 标题为空行跳过
        if not row[name_index]:
            continue

        goods_desc = row[desc_index]
        new_goods_desc = goods_desc.replace("_x000D_", "")

        goods_dict = {'id': row[id_index], 'name': row[name_index], 'desc': new_goods_desc, 'brand': row[brand_index],
                      'sku_price': row[sku_price_index], 'xy_mendian': row[xy_mendian_index], 'stock': row[stock_index],
                      "tb_time_index": tb_time_index, "xy_time_index": xy_time_index, "wd_time_index": wd_time_index,
                      "price": row[price_index], "sku_code": row[sku_code_index]}
        goods_dict_list.append(goods_dict)

        # 判断平台写入
        if row[tb_time_index] is None:
            tb_dict_list.append(goods_dict)
        if row[xy_time_index] is None:
            xy_dict_list.append(goods_dict)
        if row[wd_time_index] is None:
            wd_dict_list.append(goods_dict)

    wb.close()

    return tb_dict_list, xy_dict_list, wd_dict_list


# openpyxl 读取A列所有文本内容是 exdsf01 所在表格中的行
def read_csv_to_list(excel_path, col_str):
    # col_str 是文本内容 exdsf01
    # 准备商品数据列表
    index_list = []

    # 读取全部数据, 二维列表
    wb = load_workbook(excel_path)
    ws = wb.active

    for i, row in enumerate(ws.values):
        if row[0] == col_str:
            index_list.append(i)

    wb.close()

    return index_list


# 数据返填回表格
def write_csv(goods_id, platform, pt_dict, lm_name,excel_path):
    # 获取表格主路径


    # 准备商品数据列表
    index_list = []

    # 读取全部数据, 二维列表
    wb = load_workbook(excel_path, keep_vba=True)
    ws = wb.active

    # 获取分类和发布时间所在列
    sort_index = 0
    time_index = 0
    if platform == '淘宝':
        sort_index = pt_dict["tb_time_index"]
        time_index = sort_index + 1
    elif platform == '闲鱼':
        sort_index = pt_dict["xy_time_index"]
        time_index = sort_index + 1
    elif platform == '微店':
        sort_index = pt_dict["wd_time_index"]
        time_index = sort_index + 1
    # 获取当前日期时间
    today_date = str(datetime.date.today())

    for i, row in enumerate(ws.values):
        if row[0] == goods_id:
            index_list.append(i)
            ws.cell(row=i + 1, column=sort_index).value = lm_name
            ws.cell(row=i + 1, column=time_index).value = today_date

    time.sleep(1)
    # wb = load_workbook(excel_path, keep_vba=True)
    wb.save(excel_path)
    wb.close()
